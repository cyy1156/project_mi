#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export submission-facing Markdown (+ Excel) to PDF under 资料/初赛材料.

Usage:
  python export_all_pdfs.py
  python export_all_pdfs.py --skip-report   # skip tech-report (use md2docx_pdf.py)
  python export_all_pdfs.py --skip-excel
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]

# ---------------------------------------------------------------------------
# Minimal MD → PDF (reportlab), Chinese fonts via Windows fonts
# ---------------------------------------------------------------------------

IMG_RE = re.compile(r"^!\[([^\]]*)\]\(([^)]+)\)\s*$")
HEAD_RE = re.compile(r"^(#{1,3})\s+(.*)$")
TABLE_SEP_RE = re.compile(r"^\|?\s*:?-{3,}")


def _fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    windir = Path(r"C:\Windows\Fonts")
    candidates = [
        ("CN", windir / "msyh.ttc"),
        ("CN", windir / "msyh.ttf"),
        ("CN", windir / "simsun.ttc"),
        ("CNB", windir / "msyhbd.ttc"),
        ("CNB", windir / "simhei.ttf"),
    ]
    registered = {}
    for name, path in candidates:
        if name in registered:
            continue
        if path.exists():
            try:
                pdfmetrics.registerFont(TTFont(name, str(path), subfontIndex=0))
                registered[name] = path.name
            except Exception:
                continue
    if "CN" not in registered:
        raise RuntimeError("No Chinese TTF found under C:\\Windows\\Fonts")
    if "CNB" not in registered:
        registered["CNB"] = registered["CN"]
        pdfmetrics.registerFont(TTFont("CNB", str(windir / registered["CN"]), subfontIndex=0))
    return "CN", "CNB"


def _strip_md_inline(s: str) -> str:
    s = re.sub(r"!\[[^\]]*\]\([^)]+\)", "", s)
    s = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", s)
    s = s.replace("**", "").replace("__", "").replace("`", "")
    s = s.replace("$", "")
    return s.strip()


def md_to_pdf(md_path: Path, pdf_path: Path, title: str | None = None) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        Preformatted,
        KeepTogether,
        PageBreak,
    )
    from reportlab.lib import colors

    cn, cnb = _fonts()
    text = md_path.read_text(encoding="utf-8")
    lines = text.splitlines()

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="H1CN", fontName=cnb, fontSize=16, leading=22, spaceAfter=10, spaceBefore=14
        )
    )
    styles.add(
        ParagraphStyle(
            name="H2CN", fontName=cnb, fontSize=13, leading=18, spaceAfter=8, spaceBefore=12
        )
    )
    styles.add(
        ParagraphStyle(
            name="H3CN", fontName=cnb, fontSize=11.5, leading=16, spaceAfter=6, spaceBefore=8
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodyCN", fontName=cn, fontSize=10, leading=15, spaceAfter=4, firstLineIndent=0
        )
    )
    styles.add(
        ParagraphStyle(name="CodeCN", fontName=cn, fontSize=8.5, leading=11, backColor=colors.Color(0.95, 0.95, 0.95))
    )
    styles.add(
        ParagraphStyle(name="CellCN", fontName=cn, fontSize=8, leading=11)
    )

    story = []
    if title:
        story.append(Paragraph(_strip_md_inline(title), styles["H1CN"]))
        story.append(Spacer(1, 4 * mm))

    i = 0
    in_code = False
    code_buf: list[str] = []
    while i < len(lines):
        line = lines[i]
        if line.strip().startswith("```"):
            if not in_code:
                in_code = True
                code_buf = []
            else:
                in_code = False
                block = "\n".join(code_buf) if code_buf else " "
                story.append(Preformatted(block[:8000], styles["CodeCN"]))
                story.append(Spacer(1, 2 * mm))
            i += 1
            continue
        if in_code:
            code_buf.append(line)
            i += 1
            continue

        m = HEAD_RE.match(line)
        if m:
            level = len(m.group(1))
            body = _strip_md_inline(m.group(2))
            key = {1: "H1CN", 2: "H2CN", 3: "H3CN"}.get(level, "H3CN")
            story.append(Paragraph(body, styles[key]))
            i += 1
            continue

        if line.strip().startswith("|") and i + 1 < len(lines) and TABLE_SEP_RE.search(lines[i + 1] or ""):
            rows = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                raw = lines[i].strip().strip("|")
                if TABLE_SEP_RE.search(lines[i]):
                    i += 1
                    continue
                cells = [_strip_md_inline(c) for c in raw.split("|")]
                rows.append(cells)
                i += 1
            if rows:
                ncol = max(len(r) for r in rows)
                norm = [r + [""] * (ncol - len(r)) for r in rows]
                data = [[Paragraph(c, styles["CellCN"]) for c in r] for r in norm]
                tw = 170 * mm
                col_w = [tw / ncol] * ncol
                t = Table(data, colWidths=col_w, repeatRows=1)
                t.setStyle(
                    TableStyle(
                        [
                            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.92, 0.92, 0.95)),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 3),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                            ("TOPPADDING", (0, 0), (-1, -1), 2),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                        ]
                    )
                )
                story.append(t)
                story.append(Spacer(1, 3 * mm))
            continue

        if not line.strip():
            story.append(Spacer(1, 1.5 * mm))
            i += 1
            continue

        if IMG_RE.match(line.strip()):
            story.append(Paragraph(_strip_md_inline(line), styles["BodyCN"]))
            i += 1
            continue

        story.append(Paragraph(_strip_md_inline(line), styles["BodyCN"]))
        i += 1

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=title or md_path.stem,
    )
    doc.build(story)
    print(f"[pdf] {pdf_path} ({pdf_path.stat().st_size} bytes)")


def excel_to_pdf(xlsx: Path, pdf_path: Path) -> bool:
    """Prefer LibreOffice; fallback: dump first sheets via openpyxl+reportlab."""
    import shutil
    import subprocess

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        out_dir = pdf_path.parent
        cmd = [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(out_dir),
            str(xlsx),
        ]
        try:
            subprocess.run(cmd, check=True, capture_output=True, timeout=180)
            produced = out_dir / (xlsx.stem + ".pdf")
            if produced.exists() and produced != pdf_path:
                produced.replace(pdf_path)
            if pdf_path.exists():
                print(f"[pdf-excel] {pdf_path} via LibreOffice")
                return True
        except Exception as exc:
            print(f"[pdf-excel] LibreOffice failed: {exc}")

    try:
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.lib import colors

        cn, cnb = _fonts()
        wb = load_workbook(xlsx, data_only=True, read_only=True)
        styles_cell = ParagraphStyle(name="XC", fontName=cn, fontSize=7, leading=9)
        styles_h = ParagraphStyle(name="XH", fontName=cnb, fontSize=11, leading=14, spaceAfter=6)

        story = []
        for si, name in enumerate(wb.sheetnames[:10]):
            ws = wb[name]
            story.append(Paragraph(f"工作表：{name}", styles_h))
            rows = []
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri > 80:
                    break
                cells = []
                for v in list(row)[:12]:
                    if v is None:
                        cells.append("")
                    else:
                        s = str(v)
                        cells.append(s[:80])
                if any(c.strip() for c in cells):
                    rows.append([Paragraph(c, styles_cell) for c in cells])
            if rows:
                ncol = max(len(r) for r in rows)
                for r in rows:
                    while len(r) < ncol:
                        r.append(Paragraph("", styles_cell))
                tw = 260 * mm
                t = Table(rows, colWidths=[tw / ncol] * ncol, repeatRows=1)
                t.setStyle(
                    TableStyle(
                        [
                            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.92, 0.92, 0.95)),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ]
                    )
                )
                story.append(t)
            story.append(Spacer(1, 6 * mm))
        wb.close()
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
            title=xlsx.stem,
        )
        doc.build(story)
        print(f"[pdf-excel] {pdf_path} via openpyxl fallback")
        return True
    except Exception as exc:
        print(f"[pdf-excel] FAILED {xlsx}: {exc}")
        return False


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--skip-report", action="store_true")
    ap.add_argument("--skip-excel", action="store_true")
    args = ap.parse_args()

    jobs = [
        (ROOT / "03_演示视频" / "演示视频脚本_分镜与解说.md",
         ROOT / "03_演示视频" / "演示视频脚本_分镜与解说.pdf",
         "在线系统演示视频 · 分镜与解说词（XH-202610）"),
        (ROOT / "04_代码包" / "README_代码包说明.md",
         ROOT / "04_代码包" / "交稿" / "README_代码包说明.pdf",
         "源代码包说明（XH-202610）"),
        (ROOT / "04_代码包" / "附录A_实验证据链索引.md",
         ROOT / "04_代码包" / "交稿" / "02_附录A_实验证据链索引.pdf",
         "附录A · 实验证据链索引"),
        (ROOT / "04_代码包" / "附录B_自采数据质控样例.md",
         ROOT / "04_代码包" / "交稿" / "03_附录B_自采数据质控样例.pdf",
         "附录B · 自采数据质控样例"),
        (ROOT / "04_代码包" / "04_附录C_算法公式与算法详细介绍.md",
         ROOT / "04_代码包" / "交稿" / "04_附录C_算法公式与算法详细介绍.pdf",
         "附录C · 算法公式与算法详细介绍"),
        (ROOT / "04_代码包" / "README_采集软件说明.md",
         ROOT / "04_代码包" / "交稿" / "README_采集软件说明.pdf",
         "采集软件说明"),
        (ROOT / "02_离线验证" / "数据集使用说明.md",
         ROOT / "02_离线验证" / "交稿" / "数据集使用说明.pdf",
         "指定集数据集使用说明"),
        (ROOT / "README_初赛材料总览.md",
         ROOT / "_build" / "out" / "README_初赛材料总览.pdf",
         "初赛材料总览（内部）"),
    ]

    # Also copy appendix A md into 交稿 numbered name if missing
    app_a_src = ROOT / "04_代码包" / "附录A_实验证据链索引.md"
    app_a_dst = ROOT / "04_代码包" / "交稿" / "02_附录A_实验证据链索引.md"
    if app_a_src.exists():
        app_a_dst.write_text(app_a_src.read_text(encoding="utf-8"), encoding="utf-8")

    for md, pdf, title in jobs:
        if not md.exists():
            print(f"[skip] missing {md}")
            continue
        try:
            md_to_pdf(md, pdf, title=title)
        except Exception as exc:
            print(f"[FAIL] {md}: {exc}")

    if not args.skip_excel:
        xlsx = ROOT / "02_离线验证" / "交稿" / "离线性能验证报告_XH-202610.xlsx"
        if not xlsx.exists():
            xlsx = ROOT / "02_离线验证" / "离线性能验证报告_XH-202610.xlsx"
        if xlsx.exists():
            excel_to_pdf(xlsx, ROOT / "02_离线验证" / "交稿" / "离线性能验证报告_XH-202610.pdf")
        else:
            print(f"[skip] excel missing")

    if not args.skip_report:
        report_script = Path(__file__).resolve().parent / "md2docx_pdf.py"
        if report_script.exists():
            import runpy

            print("[report] running md2docx_pdf.py …")
            sys.argv = [str(report_script)]
            runpy.run_path(str(report_script), run_name="__main__")
        else:
            print("[report] md2docx_pdf.py not found")

    print("[done] export_all_pdfs")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
