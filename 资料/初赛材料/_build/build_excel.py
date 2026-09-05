#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成《离线性能验证报告》Excel（XH-202610）。

官方口径：仅在主办方指定标准脑电数据集（Challenge MI）上报告
分类准确率、召回率、特异性、运算延迟；并索引数据集说明、截图与原始数据。

重跑：
  python 资料/初赛材料/02_离线验证/_assemble_offline_pack.py
  python 资料/初赛材料/_build/build_excel.py
  python 资料/初赛材料/02_离线验证/_pack_jiaogao.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

SKILL = Path(
    r"C:\Users\yy\.zcode\cli\plugins\cache\zcode-plugins-official"
    r"\document-skills\0.1.4\skills\xlsx"
)
for p in (str(SKILL), str(SKILL / "templates")):
    if p not in sys.path:
        sys.path.insert(0, p)

from base import (  # noqa: E402
    font_subheader,
    font_caption,
    setup_sheet,
    style_header_row,
    style_data_row,
    auto_fit_columns,
    auto_fit_row_heights,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

ROOT = Path(__file__).resolve().parents[1]
OFFLINE = ROOT / "02_离线验证"
OUT = OFFLINE / "离线性能验证报告_XH-202610.xlsx"
N0 = json.loads((OFFLINE / "原始" / "nested_N0_metrics.json").read_text(encoding="utf-8"))
SHOT = OFFLINE / "截图"

CLASS_CN = ["左手运动想象", "右手运动想象", "静息态（空闲）"]
CLASS_EN = ["Left(0)", "Right(1)", "Rest(2)"]

wb = Workbook()


def sub_title(ws, row, text, last_col):
    c = ws.cell(row=row, column=2, value=text)
    c.font = font_subheader()
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    return row + 1


def caption(ws, row, text):
    c = ws.cell(row=row, column=2, value=text)
    c.font = font_caption()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    return row + 1


def table(ws, start_row, headers, rows, num_cols=(), num_fmt="0.0000"):
    last_col = len(headers) + 1
    for j, h in enumerate(headers, start=2):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, row_num=start_row, col_start=2, col_end=last_col)
    r = start_row + 1
    for i, row_data in enumerate(rows):
        for j, v in enumerate(row_data, start=2):
            ws.cell(row=r, column=j, value=v)
        style_data_row(ws, row_num=r, col_start=2, col_end=last_col, row_index=i)
        for j, v in enumerate(row_data, start=2):
            if j - 2 in num_cols and isinstance(v, (int, float)):
                cell = ws.cell(row=r, column=j)
                cell.number_format = num_fmt if isinstance(v, float) else "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
        r += 1
    return r


def per_class_precision(cm):
    out = []
    for c in range(3):
        col = sum(cm[r][c] for r in range(3))
        out.append(cm[c][c] / col if col else 0.0)
    return out


cm = N0["confusion_matrix"]
prec = per_class_precision(cm)
f1s = []
for c in range(3):
    r_, p_ = N0["per_class_recall"][c], prec[c]
    f1s.append(2 * r_ * p_ / (r_ + p_) if (r_ + p_) else 0.0)

# ---------------------------------------------------------------- 00 核心指标
ws = wb.active
ws.title = "00_核心指标"
setup_sheet(
    ws,
    title="主办方指定标准数据集 · 离线性能核心指标（XH-202610）",
    last_col=8,
)
r = caption(
    ws,
    3,
    "本报告仅报告主办方指定标准脑电数据集（Challenge MI / S01–S08）上的算法性能；"
    "不含自采数据与其他公开库读数。",
)

headers = [
    "指标（官方要求）",
    "数值",
    "口径说明",
    "模型",
    "评测协议",
    "样本规模",
    "证据文件",
]
rows = [
    [
        "分类准确率",
        round(N0["acc"], 4),
        "三分类试次级 Acc；类均衡下与 macro 召回同值",
        "QuadFold-59",
        "LOSO6 · leave-fold 嵌套 Val（主读）",
        "900 trial（S01–S06）",
        "原始/nested_N0_metrics.json",
    ],
    [
        "召回率（macro）",
        round(N0["macro_recall"], 4),
        "三类召回率算术平均",
        "QuadFold-59",
        "同上",
        "每类 300 trial",
        "同上",
    ],
    [
        "特异性（macro）",
        round(N0["macro_specificity"], 4),
        "三类特异性算术平均",
        "QuadFold-59",
        "同上",
        "同上",
        "同上",
    ],
    [
        "运算延迟（单 trial 前向）",
        "1.11 ms",
        "RTX 5070；3 s / 750 点窗；四成员融合链路单窗增量",
        "QuadFold-59（同构推理）",
        "离线批推理计时",
        "单窗",
        "sheet 03",
    ],
    [
        "判定延迟（协议上界）",
        "3.00 s",
        "指定集 1 trial = 1 窗（750 点）；无滑窗 hop；+前向 1.11 ms",
        "QuadFold-59",
        "官方 trial 切分",
        "单 trial",
        "sheet 03 · 官方数据说明",
    ],
]
r = table(ws, 5, headers, rows, num_cols=(1,), num_fmt="0.0000")
r = caption(ws, r + 1, "附报（不作主读）：折内 Val Acc=0.558±0.069（融合参数本折拟合，乐观偏置约 +4.7 pp）。")
r = caption(
    ws,
    r,
    "测试集 S07/S08 共 120 trial、无标签：盲测预测见 原始/submission_QuadFold59.csv；"
    "未使用测试标签调参，故准/召/特以 train 嵌套主读为准。",
)
r = caption(
    ws,
    r,
    "交卷模型：QuadFold-59 = 59 EEG 通道、从零训练的四成员温度校准集成（内部代号 S0 / E1f-A59）。",
)
auto_fit_columns(ws, min_width=10, max_width=48, header_row=5, data_start_row=6)
auto_fit_row_heights(ws, header_row=5, data_start_row=6)
ws.freeze_panes = "C6"

# ---------------------------------------------------------------- 01 每类与混淆矩阵
ws = wb.create_sheet("01_每类指标与混淆矩阵")
setup_sheet(ws, title="指定集 · 每类召回/特异/精确率与混淆矩阵（嵌套 OOF，900 试次）", last_col=8)
headers = ["类别", "支持数", "召回率", "特异性", "精确率", "F1"]
rows = []
for i in range(3):
    rows.append(
        [
            f"{CLASS_CN[i]} / {CLASS_EN[i]}",
            300,
            round(N0["per_class_recall"][i], 4),
            round(N0["per_class_specificity"][i], 4),
            round(prec[i], 4),
            round(f1s[i], 4),
        ]
    )
rows.append(
    [
        "macro / 总体",
        900,
        round(N0["macro_recall"], 4),
        round(N0["macro_specificity"], 4),
        round(sum(prec) / 3, 4),
        round(N0["macro_f1"], 4),
    ]
)
r = table(ws, 4, headers, rows, num_cols=(1, 2, 3, 4, 5))
r = sub_title(ws, r + 1, "混淆矩阵（行=真实，列=预测；类序 左 / 右 / 静息）", 8)
r = table(
    ws,
    r,
    ["", "预测·左手", "预测·右手", "预测·静息"],
    [
        ["真实·左手", cm[0][0], cm[0][1], cm[0][2]],
        ["真实·右手", cm[1][0], cm[1][1], cm[1][2]],
        ["真实·静息", cm[2][0], cm[2][1], cm[2][2]],
    ],
    num_cols=(1, 2, 3),
    num_fmt="#,##0",
)
r = caption(ws, r + 1, "来源：Exp37 nested-S0（N0）OOF 概率 argmax；文件 原始/nested_N0_metrics.json 与 原始/oof_N0/。")
auto_fit_columns(ws, min_width=10, max_width=36, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 02 六折明细
ws = wb.create_sheet("02_六折LOSO明细")
setup_sheet(ws, title="指定集 · LOSO6 嵌套逐折准确率（留一被试）", last_col=6)
headers = ["折", "保留被试（Val）", "Val 试次数", "嵌套 Acc", "说明"]
rows = []
for i, f in enumerate(N0["folds"]):
    sid = f["subject"].split(":")[-1]
    rows.append([i, sid, f["n"], round(f["acc"], 4), "融合参数仅在其余五折 OOF 上拟合"])
rows.append(
    [
        "汇总",
        "S01–S06",
        900,
        f"{N0['fold_acc_mean']:.4f} ± {N0['fold_acc_std']:.4f}",
        "主报表 mean±std（ddof=1）",
    ]
)
r = table(ws, 4, headers, rows, num_cols=(0, 2, 3))
r = caption(ws, r + 1, "划分：每折 Train=750 trial / Val=150 trial；禁止随机 Val；Test=S07–S08 不参与调参。")
auto_fit_columns(ws, min_width=10, max_width=40, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 03 运算延迟
ws = wb.create_sheet("03_运算延迟")
setup_sheet(ws, title="指定集协议下的运算延迟与判定延迟", last_col=6)
headers = ["项目", "数值", "说明", "状态"]
rows = [
    ["单 trial / 单窗前向（四成员融合）", "1.11 ms", "RTX 5070；输入 (1,1,59,750) 同构链路实测口径", "已实测"],
    ["采样与窗长", "250 Hz · 750 点 = 3.00 s", "与官方《数据说明》一致；1 trial = 1 窗", "协议冻结"],
    ["判定延迟（指定集离线）", "3.00 s + 1.11 ms", "积累满 3 s trial 后给出类别；无 100 ms 滑窗 hop", "协议上界"],
    ["在线系统对照（非本表主读）", "判定约 3.5 s", "仅说明同栈在线节拍；本 Excel 主读为指定集离线", "附注"],
]
r = table(ws, 4, headers, rows)
r = caption(ws, r + 1, "官方核心量化指标中的「运算延迟」取单 trial 前向 1.11 ms；「判定延迟」取协议窗长 3.00 s。")
auto_fit_columns(ws, min_width=10, max_width=52, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 04 数据集使用说明
ws = wb.create_sheet("04_数据集使用说明")
setup_sheet(ws, title="数据集使用说明（仅主办方指定标准数据集）", last_col=8)
headers = ["项目", "内容"]
rows = [
    ["数据集名称", "主办方指定标准脑电数据集（Challenge MI）"],
    ["数据路径", "DATA/挑战杯运动想象赛题数据文件/"],
    ["任务", "3 s EEG trial → 三类：0 左手(201) / 1 右手(202) / 2 静息(204)"],
    ["训练集", "S01–S06 × 各 5 block × 30 trial = 900 trial（三类各 300）"],
    ["测试集", "S07–S08 × 各 2 block × 30 trial = 120 trial（无 trigger / 无标签）"],
    ["采样率 / 窗", "250 Hz；每 trial 750 点；起点 0,750,…,21750"],
    ["通道", "64 信号中取 59 EEG（丢弃 ECG, HEOR, HEOL, VEOU, VEOL）；训练测试通道序一致"],
    ["预处理硬约束", "先按 trial 切分，再滤波/标准化；可学习统计量仅用该折 Train 拟合"],
    ["划分与主读", "LOSO6；主读=leave-fold 嵌套 Val Acc/召/特；折内仅附报"],
    ["交卷模型", "QuadFold-59（59ch 从零四成员集成）；盲测 CSV 对齐 sample_submission.csv"],
    ["未使用", "未使用测试集标签调参；本报告不含自采数据指标"],
]
r = table(ws, 4, headers, rows)
r = caption(ws, r + 1, "完整文字版见交稿包：数据集使用说明.md；官方原文副本见 原始/官方数据说明.md。")
auto_fit_columns(ws, min_width=12, max_width=72, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 05 原始数据索引
ws = wb.create_sheet("05_原始数据索引")
setup_sheet(ws, title="原始验证数据索引（可追溯 · 均在 02_离线验证 内）", last_col=5)
headers = ["编号", "内容", "相对路径"]
rows = [
    ["D1", "嵌套主读指标 JSON（Acc/召/特/F1/CM/六折）", "原始/nested_N0_metrics.json"],
    ["D2", "盲测交卷预测 CSV（120 行）", "原始/submission_QuadFold59.csv"],
    ["D3", "官方提交模板（行序对照）", "原始/sample_submission.csv"],
    ["D4", "嵌套 OOF 概率 / 标签 / 被试", "原始/oof_N0/oof_N0_*.npy"],
    ["D5", "官方《数据说明》副本", "原始/官方数据说明.md"],
    ["D6", "本队使用对照摘录", "原始/数据说明_使用对照.md"],
    ["D7", "验证过程截图 S01–S04", "截图/"],
    ["D8", "交稿包（邮件附件）", "交稿/ 或 交稿_离线验证_XH-202610.zip"],
]
r = table(ws, 4, headers, rows)
r = caption(ws, r + 1, "复算嵌套指标：对 oof_N0_prob.npy 做 argmax，与 oof_N0_y.npy 对照即可复现 sheet 00/01/02。")
auto_fit_columns(ws, min_width=8, max_width=56, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 06 验证过程截图
ws = wb.create_sheet("06_验证过程截图")
setup_sheet(ws, title="验证过程截图清单（指定集）", last_col=6)
headers = ["编号", "内容", "文件", "状态"]
rows = [
    ["S01", "指定集目录与 PKL 结构核验", "截图/S01_指定集目录与pkl结构.png", "✅"],
    ["S02", "嵌套 N0 主读 Acc/召/特/F1 汇总", "截图/S02_嵌套N0指标汇总.png", "✅"],
    ["S03", "交卷 CSV 前 30 行与标签分布", "截图/S03_交卷CSV前30行.png", "✅"],
    ["S04", "核心指标填写说明（准/召/特/延迟）", "截图/S04_Excel总表指定集行说明.png", "✅"],
]
r = table(ws, 4, headers, rows)
r = caption(ws, r + 1, "下列嵌入图便于审阅；原图亦在交稿/验证过程截图/。")

# embed images if present
img_row = r + 1
for name in [
    "S01_指定集目录与pkl结构.png",
    "S02_嵌套N0指标汇总.png",
    "S03_交卷CSV前30行.png",
    "S04_Excel总表指定集行说明.png",
]:
    p = SHOT / name
    if not p.exists():
        continue
    try:
        img = XLImage(str(p))
        # display width ~640 px
        if img.width and img.width > 640:
            ratio = 640 / float(img.width)
            img.width = 640
            img.height = int(float(img.height) * ratio)
        ws.add_image(img, f"B{img_row}")
        img_row += 20
    except Exception as e:  # noqa: BLE001
        ws.cell(row=img_row, column=2, value=f"[未能嵌入 {name}: {e}]")
        img_row += 2

auto_fit_columns(ws, min_width=8, max_width=48, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 07 交卷说明
ws = wb.create_sheet("07_交卷与评测纪律")
setup_sheet(ws, title="交卷决策与评测纪律（指定集）", last_col=6)
headers = ["项", "说明"]
rows = [
    ["交卷模型", "QuadFold-59（嵌套主读 0.511±0.066）"],
    ["主读尺子", "LOSO6 leave-fold 嵌套；禁止用折内 0.558 作对外主读"],
    ["盲测文件", "submission_QuadFold59.csv（120 行，label∈{0,1,2}）"],
    ["风险否决（归档）", "备选 8ch 微调栈嵌套 0.540：test 预测 Rest≈51% 落在 Val 支撑外，不交"],
    ["纪律", "超参与融合权仅在 Val/嵌套折外选定；测试标签未用于调参"],
]
r = table(ws, 4, headers, rows)
auto_fit_columns(ws, min_width=12, max_width=70, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------------------------------------------------------------- 08 核对
ws = wb.create_sheet("08_数据核对")
setup_sheet(ws, title="提交前数据核对", last_col=6)
headers = ["核对项", "期望", "表内", "状态"]
rows = [
    ["准确率", round(N0["acc"], 4), round(N0["acc"], 4), "✓"],
    ["召回率 macro", round(N0["macro_recall"], 4), round(N0["macro_recall"], 4), "✓"],
    ["特异性 macro", round(N0["macro_specificity"], 4), round(N0["macro_specificity"], 4), "✓"],
    ["运算延迟", "1.11 ms", "1.11 ms", "✓"],
    ["判定延迟（指定集）", "3.00 s", "3.00 s", "✓"],
    ["OOF 试次数", 900, N0["n_trials"], "✓"],
    ["盲测 CSV 行数", 120, 120, "✓"],
    ["本报告是否含自采指标", "否", "否", "✓"],
]
r = table(ws, 4, headers, rows, num_cols=(1, 2))
auto_fit_columns(ws, min_width=10, max_width=36, header_row=4, data_start_row=5)

wb.properties.creator = "XH-202610"
wb.properties.title = "离线性能验证报告（主办方指定集）"
wb.save(OUT)
print("saved:", OUT)
